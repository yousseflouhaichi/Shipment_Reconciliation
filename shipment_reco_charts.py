import pandas as pd
import os
import sys
import datetime
import openpyxl
from openpyxl.workbook import Workbook
from io import BytesIO

def reconcile(ship_instrutions, warehouse_reports, inv_ledger):

	data_to_excel = pd.ExcelWriter('temp/shipment_reco.xlsx')


	#input_path = 'C:\\Users\\amitu\\OneDrive\\quantuitix\\projects\\reconcify\\poc\\nmkcdd\\jillamy\\input_files'

	# booking = pd.DataFrame()
	# booking_folder = os.listdir(input_path + '\\booking_instructions')
	# for single_folder in booking_folder:
	# 	booking_files = os.listdir(input_path + '\\booking_instructions\\' + single_folder)
	# 	for single_booking in booking_files:
	#df = pd.read_excel(ship_instrutions)
	booking = pd.DataFrame()
	for ship_inst in ship_instrutions:
		ship_inst_name = ship_inst.name.split('/')[-1]
		# print(ship_inst_name)
		df = pd.read_excel(ship_inst) #####
		# print(df)
		# print(str(ship_inst_name.split('-')[0]))
		# sys.exit()
		df['BOOKING DATE'] = str(ship_inst_name.split('-')[0])
		df['BOOKING DATE'] = pd.to_datetime(df['BOOKING DATE'], format='%Y_%m_%d')
		#print(df['BOOKING DATE'])
		df.columns = df.columns.str.strip()
		df.columns = df.columns.str.upper()
		df = df[['FBA ID', 'SKU', 'CARTONS', 'QTY','BOOKING DATE']]
		df['FBA ID'] = df['FBA ID'].astype(str)
		df = df[df['FBA ID'].str.len() == 12]
		df['SKU'] = df['SKU'].astype(str)
		df['SKU'] = df['SKU'].replace(regex=True, to_replace=r'(\.0$)', value=r'')
		df['SKU'] = df['SKU'].str.replace('_New', '')
		df['SKU'] = df['SKU'].str.replace('_NEW', '')			
		df['CARTONS'] = df['CARTONS'].astype(int)
		df['QTY'] = df['QTY'].astype(int)
		#booking = booking.append(df)
		booking = booking.append(df)
		print(booking.columns)

	fba_list = list(set(booking['FBA ID'].to_list()))
	dispatch_filenames = [f + '_ViewTransaction.xlsx' for f in fba_list]

	dispatch = pd.DataFrame()
	#dispatch_files = os.listdir(warehouse_reports)
	#for single_dispatch in os.listdir(warehouse_reports):
	for single_dispatch in warehouse_reports:
		dispatch_file_name = single_dispatch.name.split('/')[-1]
		print(dispatch_file_name)
		#single_file = str(warehouse_reports +'/'+ single_dispatch)
		if dispatch_file_name in dispatch_filenames:
			#wb = openpyxl.load_workbook(single_file)
			wb = openpyxl.load_workbook(filename=BytesIO(single_dispatch.read()))
			ws1 = wb['ViewTransaction']
			fba_id = ws1['R9'].value
			dispatch_date = ws1['AD4'].value
			dispatch_date = dispatch_date.date()

			df = pd.read_excel(single_dispatch, skiprows=45, usecols='C:AS')
			df['FBA ID'] = fba_id
			df['DISPATCH DATE'] = dispatch_date
			df['DISPATCH DATE'] = pd.to_datetime(df['DISPATCH DATE'], format='%Y-%m-%d')
			df.columns = df.columns.str.strip()
			df.columns = df.columns.str.upper()
			df = df[['FBA ID', 'DISPATCH DATE', 'SKU', 'INV QTY']]
			df.dropna(subset=['SKU'], inplace=True)
			remove_cols = [col for col in df.columns if 'Unnamed' in col]
			df.drop(remove_cols, axis=1, inplace=True)
			df['SKU'] = df['SKU'].astype(str)
			df['SKU'] = df['SKU'].replace(regex=True, to_replace=r'(\.0$)', value=r'')
			df['INV QTY'] = df['INV QTY'].astype(int)

			dispatch = dispatch.append(df)
			#print('######################')
			#print(dispatch.head())
	#print(dispatch.columns)

	#inventory = pd.read_csv(inv_ledger)
	inventory = inv_ledger
	inventory['Date'] = pd.to_datetime(inventory['Date'], format='%m/%d/%Y')
	inventory['MSKU'] = inventory['MSKU'].astype(str).str[0:12]
	inventory_receipts = inventory[inventory['Event Type'] == 'Receipts'].rename(columns={'Reference ID': 'FBA ID', 'MSKU': 'SKU'})

	inventory_extract = pd.DataFrame()
	for single_fba in fba_list:
		df = inventory_receipts[inventory_receipts['FBA ID'] == single_fba]
		df.columns = df.columns.str.strip()
		df.columns = df.columns.str.upper()
		df = df[['FBA ID', 'DATE', 'SKU', 'QUANTITY']].rename(columns={'DATE': 'RECEIPT DATE'})
		# df = df[['']]
		inventory_extract = inventory_extract.append(df)
	#print(inventory_extract)
	print(booking.info())
	print(dispatch.info())

	fba_detail = pd.merge(booking, dispatch, on=['FBA ID', 'SKU'], how='outer')
	fba_detail = pd.merge(fba_detail, inventory_extract, on=['FBA ID', 'SKU'], how='outer')
	fba_detail = fba_detail.rename(columns={'CARTONS': 'CARTONS BOOKED', 'INV QTY': 'CARTONS DISPATCHED', 'QTY': 'UNITS BOOKED', 'QUANTITY': 'UNITS RECEIVED'})

	fba_detail['DISPATCH DAYS'] = fba_detail['DISPATCH DATE'] - fba_detail['BOOKING DATE']
	fba_detail['RECEIPT DAYS'] = fba_detail['RECEIPT DATE'] - fba_detail['DISPATCH DATE']

	fba_detail['BOOKING DATE'] = fba_detail['BOOKING DATE'].dt.strftime('%Y-%m-%d')
	fba_detail['DISPATCH DATE'] = fba_detail['DISPATCH DATE'].dt.strftime('%Y-%m-%d')
	fba_detail['RECEIPT DATE'] = fba_detail['RECEIPT DATE'].dt.strftime('%Y-%m-%d')

	fba_detail['BOOKING DATE'].fillna('NOT AVAILABLE', inplace=True)
	fba_detail['DISPATCH DATE'].fillna('NOT AVAILABLE', inplace=True)
	fba_detail['CARTONS BOOKED'].fillna(0, inplace=True)
	fba_detail['CARTONS DISPATCHED'].fillna(0, inplace=True)
	fba_detail['UNITS BOOKED'].fillna(0, inplace=True)
	fba_detail['UNITS RECEIVED'].fillna(0, inplace=True)
	fba_detail['RECEIPT DATE'].fillna('NOT AVAILABLE', inplace=True)
	fba_detail['DISPATCH DAYS'].fillna(datetime.timedelta(days=0), inplace=True)
	fba_detail['RECEIPT DAYS'].fillna(datetime.timedelta(days=0), inplace=True)
	fba_detail.sort_values(by=['FBA ID', 'BOOKING DATE', 'DISPATCH DATE', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED'], inplace=True)

	fba_detail.set_index(['FBA ID', 'BOOKING DATE', 'DISPATCH DATE', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED'], inplace=True)
	#print(fba_detail.info())
	# fba_detail.to_excel('fba_detail.xlsx')
	# sys.exit()

	fba_skuwise = fba_detail.reset_index()
	fba_skuwise = fba_skuwise.groupby(['FBA ID', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'UNITS BOOKED']).agg({'UNITS RECEIVED': 'sum'}).reset_index()
	fba_skuwise['CARTONS SHORT DISPATCHED'] = fba_skuwise['CARTONS BOOKED'] - fba_skuwise['CARTONS DISPATCHED']
	fba_skuwise['UNITS SHORT RECEIVED'] = fba_skuwise['UNITS BOOKED'] - fba_skuwise['UNITS RECEIVED']
	fba_skuwise.set_index(['FBA ID', 'SKU'], inplace=True)
	fba_skuwise = fba_skuwise[['CARTONS BOOKED', 'CARTONS DISPATCHED', 'CARTONS SHORT DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED', 'UNITS SHORT RECEIVED']]
	#print(fba_skuwise.info())
	# sys.exit()

	fba_exception = fba_skuwise[(fba_skuwise['CARTONS SHORT DISPATCHED'] != 0) | (fba_skuwise['UNITS SHORT RECEIVED'] != 0)]
	#print(fba_exception.info())

	sum_positives = lambda x: x[x>0].sum()
	sum_negatives = lambda x: x[x<0].sum() * (-1)
	fba_summary = fba_skuwise.reset_index()
	fba_summary = fba_summary.rename(columns={'CARTONS SHORT DISPATCHED': 'EXCESS CARTONS DISPATCHED', 'UNITS SHORT RECEIVED': 'EXCESS UNITS RECEIVED'})
	fba_summary['SHORT CARTONS DISPATCHED'] = fba_summary['EXCESS CARTONS DISPATCHED']
	fba_summary['SHORT UNITS RECEIVED'] = fba_summary['EXCESS UNITS RECEIVED']
	fba_summary = fba_summary.groupby(['FBA ID']).agg({'SKU': 'count', 'CARTONS BOOKED': 'sum', 'CARTONS DISPATCHED': 'sum', 'UNITS BOOKED': 'sum', 'UNITS RECEIVED': 'sum', 'EXCESS CARTONS DISPATCHED': sum_negatives, 'SHORT CARTONS DISPATCHED': sum_positives, 'EXCESS UNITS RECEIVED': sum_negatives, 'SHORT UNITS RECEIVED': sum_positives}).reset_index()

	fba_summary = fba_summary[['FBA ID', 'SKU', 'CARTONS BOOKED', 'CARTONS DISPATCHED', 'EXCESS CARTONS DISPATCHED', 'SHORT CARTONS DISPATCHED', 'UNITS BOOKED', 'UNITS RECEIVED', 'EXCESS UNITS RECEIVED', 'SHORT UNITS RECEIVED']]
	#print(fba_summary.info())

	fba_summary.to_excel(data_to_excel, sheet_name='FBA Summary', index=False)
	fba_exception.to_excel(data_to_excel, sheet_name='FBA Exceptions')
	fba_skuwise.to_excel(data_to_excel, sheet_name='FBA SKU-wise')
	fba_detail.to_excel(data_to_excel, sheet_name='FBA Detail')

	#data_to_excel.save()

	workbook = data_to_excel.book
	date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
	number_format = workbook.add_format({'num_format': '#,##0'})
	fail_format = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
	pass_format = workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'})
	center_format = workbook.add_format()
	center_format.set_align('center')
	right_format = workbook.add_format()
	right_format.set_align('center')

	sheet1 = data_to_excel.sheets['FBA Summary']
	sheet2 = data_to_excel.sheets['FBA Exceptions']
	sheet3 = data_to_excel.sheets['FBA SKU-wise']
	sheet4 = data_to_excel.sheets['FBA Detail']

	# sheet1.set_column('A:A', 22, center_format)
	sheet1.set_column('B:J', 22, number_format)
	sheet2.set_column('B:B', 22, center_format)
	sheet2.set_column('C:H', 22, number_format)
	sheet3.set_column('B:B', 22, date_format)
	sheet3.set_column('C:H', 22, number_format)						
	sheet4.set_column('B:D', 22, center_format)
	sheet4.set_column('E:H', 22, number_format)
	sheet4.set_column('I:I', 22, center_format)
	sheet4.set_column('J:K', 22, number_format)

	sheet1.conditional_format('E2:F'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet1.conditional_format('E2:F'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})
	sheet1.conditional_format('I2:J'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet1.conditional_format('I2:J'+str(len(fba_summary)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet2.conditional_format('E2:E'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet2.conditional_format('E2:E'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})
	sheet2.conditional_format('H2:H'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet2.conditional_format('H2:H'+str(len(fba_exception)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet3.conditional_format('E2:E'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet3.conditional_format('E2:E'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})
	sheet3.conditional_format('H2:H'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '!=', 'value': 0, 'format': fail_format})
	sheet3.conditional_format('H2:H'+str(len(fba_skuwise)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': pass_format})

	sheet4.conditional_format('B2:B'+str(len(fba_detail)+1), {'type': 'text', 'criteria': 'containing', 'value': 'NOT AVAILABLE', 'format': fail_format})
	sheet4.conditional_format('C2:C'+str(len(fba_detail)+1), {'type': 'text', 'criteria': 'containing', 'value': 'NOT AVAILABLE', 'format': fail_format})
	sheet4.conditional_format('I2:I'+str(len(fba_detail)+1), {'type': 'text', 'criteria': 'containing', 'value': 'NOT AVAILABLE', 'format': fail_format})
	sheet4.conditional_format('E2:H'+str(len(fba_detail)+1), {'type': 'cell', 'criteria': '=', 'value': 0, 'format': fail_format})

	# sheet1.set_column(0, 6, 22)
	# sheet2.set_column(0, 5, 22)
	# sheet3.set_column(0, 5, 22)
	# sheet4.set_column(0, 6, 22)			

	data_to_excel.save()

	workbook = openpyxl.load_workbook('temp/shipment_reco.xlsx')
	#workbook = Workbook()
	sheet1 = workbook['FBA Summary']
	sheet2 = workbook['FBA Exceptions']
	sheet3 = workbook['FBA SKU-wise']
	sheet4 = workbook['FBA Detail']


	sheet1.cell(row=len(fba_summary)+2, column=2).value = fba_summary['SKU'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=3).value = fba_summary['CARTONS BOOKED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=4).value = fba_summary['CARTONS DISPATCHED'].sum()			
	sheet1.cell(row=len(fba_summary)+2, column=5).value = fba_summary['EXCESS CARTONS DISPATCHED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=6).value = fba_summary['SHORT CARTONS DISPATCHED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=7).value = fba_summary['UNITS BOOKED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=8).value = fba_summary['UNITS RECEIVED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=9).value = fba_summary['EXCESS UNITS RECEIVED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=10).value = fba_summary['SHORT UNITS RECEIVED'].sum()
	sheet1.cell(row=len(fba_summary)+2, column=2).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=3).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=6).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=7).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=8).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=9).font = openpyxl.styles.Font(bold=True)
	sheet1.cell(row=len(fba_summary)+2, column=10).font = openpyxl.styles.Font(bold=True)

	sheet2.cell(row=len(fba_exception)+2, column=5).value = fba_exception['CARTONS SHORT DISPATCHED'].sum()
	sheet2.cell(row=len(fba_exception)+2, column=8).value = fba_exception['UNITS SHORT RECEIVED'].sum()
	sheet2.cell(row=len(fba_exception)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet2.cell(row=len(fba_exception)+2, column=8).font = openpyxl.styles.Font(bold=True)

	sheet3.cell(row=len(fba_skuwise)+2, column=3).value = fba_skuwise['CARTONS BOOKED'].sum()
	sheet3.cell(row=len(fba_skuwise)+2, column=4).value = fba_skuwise['CARTONS DISPATCHED'].sum()
	sheet3.cell(row=len(fba_skuwise)+2, column=5).value = fba_skuwise['CARTONS SHORT DISPATCHED'].sum()			
	sheet3.cell(row=len(fba_skuwise)+2, column=6).value = fba_skuwise['UNITS BOOKED'].sum()
	sheet3.cell(row=len(fba_skuwise)+2, column=7).value = fba_skuwise['UNITS RECEIVED'].sum()
	sheet3.cell(row=len(fba_skuwise)+2, column=8).value = fba_skuwise['UNITS SHORT RECEIVED'].sum()
	sheet3.cell(row=len(fba_skuwise)+2, column=3).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(fba_skuwise)+2, column=4).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(fba_skuwise)+2, column=5).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(fba_skuwise)+2, column=6).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(fba_skuwise)+2, column=7).font = openpyxl.styles.Font(bold=True)
	sheet3.cell(row=len(fba_skuwise)+2, column=8).font = openpyxl.styles.Font(bold=True)

	for c in range(1, 11):
		sheet1.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
	# 	sheet1.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

	for c in range(1, 9):
		sheet2.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
	# 	sheet2.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

	for c in range(1, 9):
		sheet3.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
	# 	sheet3.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

	for c in range(1, 12):
		sheet4.cell(row=1, column=c).font = openpyxl.styles.Font(size=9, bold=True)
	# 	sheet4.cell(row=1, column=c).fill = openpyxl.styles.PatternFill(bgColor='D9D9D9', fill_type='solid')

	for c in ('A', 'B'):
		for r in range(2, len(fba_exception)+2):
			sheet2[c+str(r)].font = openpyxl.styles.Font(bold=False)

	for c in ('A', 'B'):
		for r in range(2, len(fba_skuwise)+2):
			sheet3[c+str(r)].font = openpyxl.styles.Font(bold=False)

	for c in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
		for r in range(2, len(fba_detail)+2):
			sheet4[c+str(r)].font = openpyxl.styles.Font(bold=False)

	border = openpyxl.styles.borders.Side(style = None, color = '00000000', border_style = 'thin')
	border_format = openpyxl.styles.borders.Border(left = border, right = border, bottom = border, top = border)

	for c in ('C', 'D', 'E', 'F', 'G', 'H'):
		for r in range(2, len(fba_exception)+3):
			sheet2[c+str(r)].border = border_format

	for c in ('C', 'D', 'E', 'F', 'G', 'H'):
		for r in range(2, len(fba_skuwise)+3):
			sheet3[c+str(r)].border = border_format

	for c in ('I', 'J', 'K'):
		for r in range(2, len(fba_detail)+2):
			sheet4[c+str(r)].border = border_format

	units_booked = fba_summary['UNITS BOOKED'].sum()
	excess_units_received = fba_summary['EXCESS UNITS RECEIVED'].sum()
	short_units_received = -fba_summary['SHORT UNITS RECEIVED'].sum()	
	units_received = -fba_summary['UNITS RECEIVED'].sum()
	
	matching_sku = sum(fba_skuwise['UNITS SHORT RECEIVED'] == 0)
	mismatching_sku = sum(fba_skuwise['UNITS SHORT RECEIVED'] != 0)

	workbook.save('temp/shipment_reco.xlsx')
	
	return units_booked, excess_units_received, short_units_received, units_received, matching_sku, mismatching_sku

	# data_to_excel.save()




	# Rearrange columns (done)
	# Formating
		# Colour (done)
		# Width (done)
		# Border (done)
		# Totals (done)
		# Unbold (done)
		# Header font size (done)
		# Header colour (done)
		# Cell alignment (done)
		# Number format (done)
	# Adjustment units

# ship_instrutions = 'sample_data/booking_instructions/2022_04_19/Jillamy Booking file 19th April.xlsx'
# warehouse_reports = 'sample_data/warehouse_reports'
# inv_ledger = 'sample_data/inventory_ledger/inventory_ledger.csv'

# work = reconcile(ship_instrutions, warehouse_reports, inv_ledger)